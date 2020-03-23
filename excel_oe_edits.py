import openpyxl
from openpyxl import load_workbook
from color_class import bcolors
import os
import datetime

year = datetime.date.today().year

def excel_work(projectCode,projectNumber,accountNumber):
    #_file = os.path.join(r'P:\ONLINE DATA\Online Data Export',accountNumber,projectCode,accountNumber+'_'+projectCode+'_'+'Open-End_C.xlsx')
    _file = os.path.join(r'P:\TAB\Tab Export',str(year),projectNumber,accountNumber+'_'+projectCode+'_'+projectNumber+'_'+'OE_C.xlsx')
    
    try:
        workbook = load_workbook(filename = _file)

        sheet = workbook.active
        
        #Rename Cell to Vendor
        sheet["N1"] = "Vendors"
        print('.')
        print('.')
        print('.')
        print(bcolors.OKBLUE+"'Cell' has been renamed to 'Vendors'"+bcolors.ENDC)
        print('.')
        print('.')
        print('.')

        #delete typical columns
        sheet.delete_cols(idx=2,amount=2)
        sheet.delete_cols(idx=3,amount=3)
        sheet.delete_cols(idx=4,amount=5)
        sheet.delete_cols(idx=5,amount=2)
        print('.')
        print('.')
        print('.')
        print(bcolors.OKBLUE+'Typcical Columns were deleted.'+bcolors.ENDC)
        print('.')
        print('.')
        print('.')

        #check for PII and delete
        #Month done
        #Day done
        #Year done
        #City done
        #State not in OE
        #Zip done
        #Email done

        for index, column in reversed(list(enumerate(sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=sheet.max_column,values_only=True)))):
            column_name = column[0]
            #print(index,column_name)
            
            # #birthdate
            if "-Month" in column_name:
                sheet.delete_cols(idx=index+1,amount=1)
                print(bcolors.OKBLUE+'This PII column has been deleted:'+bcolors.ENDC,column[0])

            if "-Day" in column_name:
                sheet.delete_cols(idx=index+1,amount=1)
                print(bcolors.OKBLUE+'This PII column has been deleted:'+bcolors.ENDC,column[0])

            if "-Year" in column_name:
                sheet.delete_cols(idx=index+1,amount=1)
                print(bcolors.OKBLUE+'This PII column has been deleted:'+bcolors.ENDC,column[0])

            #location
            if "-City" in column_name:
                sheet.delete_cols(idx=index+1,amount=1)
                print(bcolors.OKBLUE+'This PII column has been deleted:'+bcolors.ENDC,column[0])
            
            if "-Zip" in column_name:
                sheet.delete_cols(idx=index+1,amount=1)
                print(bcolors.OKBLUE+'This PII column has been deleted:'+bcolors.ENDC,column[0])

            #Email
            if "-E-mail" in column_name:
                sheet.delete_cols(idx=index+1,amount=1)
                print(bcolors.OKBLUE+'This PII column has been deleted:'+bcolors.ENDC,column[0])

        #save the file
        workbook.save(_file)
        return True
    except Exception as e:
        print(bcolors.FAIL+'ERROR'+bcolors.ENDC)
        print(bcolors.FAIL+str(e)+bcolors.ENDC)
        return False
        
#excelWork(r'C:\Users\DRivers\Desktop\Python\Post Export Automation\test_OE.xlsx')